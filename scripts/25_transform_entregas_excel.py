"""Transformación local y trazable de Excel de entregas por emergencias.

Lee los Excel originales y la auditoría generada por el script 24. Produce una
tabla normalizada, alertas por fila y un resumen. No modifica fuentes, no elimina
duplicados y no interactúa con dashboard, SQL ni TiDB.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable

import pandas as pd
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path(r"C:\Users\Usuario\Documents\BAUTI\Emergencias\Entregas")
AUDIT_DIR = PROJECT_DIR / "data_processed" / "entregas" / "audit"
OUTPUT_DIR = PROJECT_DIR / "data_processed" / "entregas" / "normalized"
ORIGIN = "entregas_emergencia_excel"
VALID_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}

FACT_COLUMNS = [
    "entrega_id", "anio", "fecha_entrega", "productor_nombre",
    "cuit_cuil_original", "cuit_cuil_norm", "documento_original",
    "documento_norm", "renspa", "adrema", "departamento", "localidad",
    "municipio", "paraje", "tipo_asistencia", "proveedor",
    "insumo_producto", "cantidad", "unidad", "monto_estimado", "moneda",
    "norma_evento", "expediente", "fuente_archivo", "fuente_hoja",
    "source_row_number", "origen_dato", "calidad_identificacion", "observaciones",
]

ALERT_COLUMNS = [
    "entrega_id", "fuente_archivo", "fuente_hoja", "source_row_number",
    "alerta_tipo", "alerta_descripcion", "severidad",
]

FIELD_PATTERNS: dict[str, list[str]] = {
    "productor_nombre": [r"nombre.*razon.*social", r"razon.*social", r"nombre.*apellido", r"beneficiari", r"^productor$", r"^titular$"],
    "cuit_cuil": [r"cuit.*cuil", r"cuil.*cuit", r"^cuit$", r"^cuil$"],
    "documento": [r"^dni$", r"document", r"nro.*doc", r"n.*doc"],
    "renspa": [r"^renspa$"],
    "adrema": [r"adrema"],
    "departamento": [r"departamento", r"^dpto$"],
    "localidad": [r"localidad"],
    "municipio": [r"municipio"],
    "paraje": [r"paraje"],
    "fecha_entrega": [r"fecha.*entrega", r"^fecha$"],
    "anio": [r"^anio$", r"^ano$"],
    "tipo_asistencia": [r"tipo.*asistencia", r"tipo.*entrega"],
    "proveedor": [r"^proveedor$"],
    "insumo_producto": [r"^insumo", r"^producto$", r"concepto", r"material"],
    "cantidad": [r"total.*kilos", r"cant.*rollos", r"^cantidad$", r"^cantidades$", r"^cant$", r"unidades"],
    "unidad": [r"^unidad$", r"unidad.*medida", r"kg.*bolsa"],
    "monto_estimado": [r"^monto$", r"^importe$", r"^valor$", r"^subsidio$", r"^saldo$"],
    "moneda": [r"moneda"],
    "norma_evento": [r"resolucion", r"decreto", r"^norma$", r"instrumento.*legal"],
    "expediente": [r"expediente", r"^expte$", r"^exp$"],
    "observaciones": [r"observ", r"comentario"],
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().strip()
    return re.sub(r"[^a-z0-9]+", "_", text).strip("_")


def clean_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and pd.isna(value):
        return ""
    text = re.sub(r"\s+", " ", str(value)).strip()
    if normalize_text(text) in {"", "nan", "none", "null", "n_a", "s_d", "sin_dato", "_"}:
        return ""
    return text


def clean_identifier_original(value: Any) -> str:
    text = clean_text(value)
    return re.sub(r"\.0$", "", text) if text else ""


def digits_only(value: Any) -> str:
    text = clean_identifier_original(value)
    return re.sub(r"\D", "", text)


def normalize_cuit(value: Any) -> tuple[str, str]:
    original = clean_identifier_original(value)
    digits = digits_only(original)
    return original, digits if len(digits) == 11 else ""


def normalize_document(value: Any) -> tuple[str, str]:
    original = clean_identifier_original(value)
    digits = digits_only(original)
    return original, digits


def clean_renspa(value: Any) -> str:
    text = clean_text(value)
    if not text:
        return ""
    text = re.sub(r"\s*-\s*", "-", text)
    text = re.sub(r"-{2,}", "-", text)
    return text.strip(" -")


def clean_adrema(value: Any) -> str:
    return re.sub(r"\s+", "", clean_text(value))


def parse_number(value: Any) -> tuple[float | None, bool]:
    """Devuelve número y bandera de fallo para valores no vacíos."""
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return None, False
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return float(value), False
    text = clean_text(value)
    if not text:
        return None, False
    text = re.sub(r"[^0-9,.-]", "", text)
    if not re.search(r"\d", text):
        return None, True
    negative = text.startswith("-")
    text = text.replace("-", "")
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        parts = text.split(",")
        text = "".join(parts) if len(parts[-1]) == 3 and len(parts) > 1 else "".join(parts[:-1]) + "." + parts[-1]
    elif "." in text:
        parts = text.split(".")
        if len(parts) > 2 or (len(parts[-1]) == 3 and len(parts) > 1):
            text = "".join(parts)
    try:
        number = float(text)
        return (-number if negative else number), False
    except ValueError:
        return None, True


def parse_date(value: Any) -> tuple[str, bool]:
    if value is None or (isinstance(value, float) and pd.isna(value)):
        return "", False
    if isinstance(value, (datetime, date)):
        return value.date().isoformat() if isinstance(value, datetime) else value.isoformat(), False
    if isinstance(value, (int, float)) and 20_000 <= float(value) <= 80_000:
        parsed = datetime(1899, 12, 30) + timedelta(days=float(value))
        return parsed.date().isoformat(), False
    text = clean_text(value)
    if not text:
        return "", False
    parsed = pd.to_datetime(text, dayfirst=True, errors="coerce")
    if pd.isna(parsed):
        return "", True
    return parsed.date().isoformat(), False


def bool_value(value: Any) -> bool:
    return str(value).strip().lower() in {"true", "1", "si", "sí"}


def make_headers(values: Iterable[Any]) -> list[str]:
    counts: Counter[str] = Counter()
    headers: list[str] = []
    for position, value in enumerate(values, start=1):
        base = clean_text(value) or f"sin_nombre_{position}"
        counts[base] += 1
        headers.append(base if counts[base] == 1 else f"{base}_{counts[base]}")
    return headers


def locate_columns(headers: list[str]) -> dict[str, int]:
    result: dict[str, int] = {}
    normalized = [normalize_text(header) for header in headers]
    for field, patterns in FIELD_PATTERNS.items():
        for pattern in patterns:
            for position, header in enumerate(normalized):
                if field == "renspa" and "pdf" in header:
                    continue
                if re.search(pattern, header):
                    result[field] = position
                    break
            if field in result:
                break
    return result


def value_at(row: tuple[Any, ...], mapping: dict[str, int], field: str) -> Any:
    position = mapping.get(field)
    return row[position] if position is not None and position < len(row) else None


def infer_year(column_value: Any, parsed_date: str, path: Path) -> int | None:
    numeric, failed = parse_number(column_value)
    if not failed and numeric is not None and 1900 <= int(numeric) <= 2100:
        return int(numeric)
    if parsed_date:
        return int(parsed_date[:4])
    match = re.search(r"\b(2022|2024|2025)\b", path.name)
    if match:
        return int(match.group(1))
    for part in reversed(path.parts):
        if part in {"2022", "2024", "2025"}:
            return int(part)
    return None


def infer_provider(column_value: Any, filename: str) -> str:
    value = clean_text(column_value)
    if value:
        return value
    normalized = normalize_text(filename)
    for token, label in [("prieto", "Prieto"), ("koffman", "Kofman"), ("kofman", "Kofman"), ("fitosan", "Fitosan"), ("borderes", "Borderes")]:
        if token in normalized:
            return label
    return "Sin dato"


def infer_assistance(column_value: Any, filename: str, sheetname: str) -> str:
    value = clean_text(column_value)
    if value:
        return value
    text = normalize_text(f"{filename} {sheetname}")
    if "saldo" in text and "emergencia" in text:
        return "saldo emergencia"
    if "subsid" in text:
        return "subsidio"
    if any(token in text for token in ("rollos", "kg", "kofman", "prieto", "fitosan", "borderes")):
        return "entrega de insumos"
    if "especie" in text or "asistid" in text:
        return "asistencia en especie"
    return "otro"


def infer_product(column_value: Any, filename: str, sheetname: str) -> str:
    value = clean_text(column_value)
    if value:
        return value
    text = normalize_text(f"{sheetname} {filename}")
    if "rollos" in text:
        return "Rollos plásticos"
    if "saldo_emergencia" in text:
        return "Saldo emergencia"
    # Un proveedor y una cantidad en kg no bastan para afirmar qué producto fue entregado.
    return "Sin dato"


def infer_unit(headers: list[str], mapping: dict[str, int], row: tuple[Any, ...], sheetname: str) -> str:
    explicit = clean_text(value_at(row, mapping, "unidad"))
    if explicit:
        return explicit
    quantity_position = mapping.get("cantidad")
    quantity_header = normalize_text(headers[quantity_position]) if quantity_position is not None else ""
    if "kilo" in quantity_header or re.search(r"\d[\d._]*kg", normalize_text(sheetname)):
        return "kg"
    if "rollo" in quantity_header or "rollo" in normalize_text(sheetname):
        return "rollo"
    return ""


def stable_id(record: dict[str, Any]) -> str:
    parts = [
        record["fuente_archivo"], record["fuente_hoja"], str(record["source_row_number"]),
        record["productor_nombre"], record["cuit_cuil_norm"], record["documento_norm"],
    ]
    return hashlib.sha256("|".join(parts).encode("utf-8")).hexdigest()


def identification_quality(record: dict[str, Any]) -> str:
    if len(record["cuit_cuil_norm"]) == 11:
        return "alta"
    if record["productor_nombre"] and record["documento_norm"]:
        return "media"
    if record["productor_nombre"] and (record["renspa"] or record["adrema"]):
        return "media"
    if record["productor_nombre"]:
        return "baja"
    return "pendiente"


def add_alert(alerts: list[dict[str, Any]], record: dict[str, Any], kind: str, description: str, severity: str) -> None:
    alerts.append({
        "entrega_id": record["entrega_id"], "fuente_archivo": record["fuente_archivo"],
        "fuente_hoja": record["fuente_hoja"], "source_row_number": record["source_row_number"],
        "alerta_tipo": kind, "alerta_descripcion": description, "severidad": severity,
    })


def select_sheets() -> pd.DataFrame:
    sheets = pd.read_csv(AUDIT_DIR / "entregas_excel_sheets.csv").fillna("")
    keys = pd.read_csv(AUDIT_DIR / "entregas_excel_key_fields_detected.csv").fillna("")
    merged = sheets.merge(keys, on=["nombre_archivo", "nombre_hoja"], how="inner")
    structured_id = merged[["tiene_cuit", "tiene_dni", "tiene_renspa", "tiene_adrema"]].apply(
        lambda row: any(bool_value(value) for value in row), axis=1
    )
    has_name = merged["tiene_nombre_productor"].map(bool_value)
    is_table = merged["parece_tabla_o_resumen"].eq("tabla")
    return merged[structured_id & has_name & is_table].copy()


def source_files_by_name() -> dict[str, Path]:
    return {
        path.name: path for path in SOURCE_DIR.rglob("*")
        if path.is_file() and path.suffix.lower() in VALID_EXTENSIONS and not path.name.startswith("~$")
    }


def transform() -> tuple[pd.DataFrame, pd.DataFrame, list[dict[str, Any]]]:
    selected = select_sheets()
    file_map = source_files_by_name()
    records: list[dict[str, Any]] = []
    alerts: list[dict[str, Any]] = []
    processed_sheets: list[dict[str, Any]] = []

    for filename, group in selected.groupby("nombre_archivo", sort=True):
        path = file_map.get(filename)
        if path is None:
            raise FileNotFoundError(f"No se encontró el Excel auditado: {filename}")
        workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        for _, audit_row in group.iterrows():
            sheetname = audit_row["nombre_hoja"]
            if sheetname not in workbook.sheetnames:
                raise KeyError(f"No se encontró la hoja auditada {sheetname!r} en {filename}")
            ws = workbook[sheetname]
            header_row = int(float(audit_row["primera_fila_con_datos_probable"]))
            header_values = next(ws.iter_rows(min_row=header_row, max_row=header_row, values_only=True))
            headers = make_headers(header_values)
            mapping = locate_columns(headers)
            before = len(records)

            for source_row_number, row in enumerate(
                ws.iter_rows(min_row=header_row + 1, max_row=ws.max_row, max_col=len(headers), values_only=True),
                start=header_row + 1,
            ):
                producer = clean_text(value_at(row, mapping, "productor_nombre"))
                cuit_original, cuit_norm = normalize_cuit(value_at(row, mapping, "cuit_cuil"))
                document_original, document_norm = normalize_document(value_at(row, mapping, "documento"))
                renspa = clean_renspa(value_at(row, mapping, "renspa"))
                adrema = clean_adrema(value_at(row, mapping, "adrema"))
                location = clean_text(value_at(row, mapping, "localidad"))

                # Las filas sin sujeto ni identificador son títulos, totales o espacios y no constituyen entregas.
                if not any((producer, cuit_original, document_original, renspa, adrema)):
                    continue

                date_original = value_at(row, mapping, "fecha_entrega")
                parsed_date, date_failed = parse_date(date_original)
                year = infer_year(value_at(row, mapping, "anio"), parsed_date, path)
                provider = infer_provider(value_at(row, mapping, "proveedor"), filename)
                assistance = infer_assistance(value_at(row, mapping, "tipo_asistencia"), filename, sheetname)
                product = infer_product(value_at(row, mapping, "insumo_producto"), filename, sheetname)
                quantity_original = value_at(row, mapping, "cantidad")
                quantity, quantity_failed = parse_number(quantity_original)
                amount_original = value_at(row, mapping, "monto_estimado")
                amount, amount_failed = parse_number(amount_original)
                unit = infer_unit(headers, mapping, row, sheetname)
                currency = clean_text(value_at(row, mapping, "moneda"))
                if amount is not None and not currency:
                    currency = "ARS"

                record: dict[str, Any] = {
                    "entrega_id": "", "anio": year, "fecha_entrega": parsed_date,
                    "productor_nombre": producer, "cuit_cuil_original": cuit_original,
                    "cuit_cuil_norm": cuit_norm, "documento_original": document_original,
                    "documento_norm": document_norm, "renspa": renspa, "adrema": adrema,
                    "departamento": clean_text(value_at(row, mapping, "departamento")),
                    "localidad": location, "municipio": clean_text(value_at(row, mapping, "municipio")),
                    "paraje": clean_text(value_at(row, mapping, "paraje")),
                    "tipo_asistencia": assistance, "proveedor": provider,
                    "insumo_producto": product, "cantidad": quantity, "unidad": unit,
                    "monto_estimado": amount, "moneda": currency,
                    "norma_evento": clean_text(value_at(row, mapping, "norma_evento")),
                    "expediente": clean_text(value_at(row, mapping, "expediente")),
                    "fuente_archivo": filename, "fuente_hoja": sheetname,
                    "source_row_number": source_row_number, "origen_dato": ORIGIN,
                    "calidad_identificacion": "",
                    "observaciones": clean_text(value_at(row, mapping, "observaciones")),
                }
                record["entrega_id"] = stable_id(record)
                record["calidad_identificacion"] = identification_quality(record)
                records.append(record)

                if not any((cuit_norm, document_norm, renspa, adrema, producer)):
                    add_alert(alerts, record, "sin_identificador", "No posee identificador ni nombre útil.", "alta")
                elif producer and not any((cuit_norm, document_norm, renspa, adrema)):
                    add_alert(alerts, record, "solo_nombre", "Solo dispone del nombre del productor como clave de vinculación.", "media")
                if cuit_original and not cuit_norm:
                    add_alert(alerts, record, "cuit_invalido", "El CUIT/CUIL original no contiene exactamente 11 dígitos.", "alta")
                if not document_norm:
                    add_alert(alerts, record, "documento_faltante", "No se dispone de DNI/documento normalizado.", "baja")
                if date_failed:
                    add_alert(alerts, record, "fecha_no_parseable", "La fecha informada no pudo convertirse de forma segura.", "media")
                if year is None:
                    add_alert(alerts, record, "anio_no_inferido", "No se pudo inferir el año por columna, fecha, archivo o ruta.", "alta")
                if quantity_failed:
                    add_alert(alerts, record, "cantidad_no_parseable", "La cantidad informada no pudo convertirse a número.", "media")
                if amount_failed:
                    add_alert(alerts, record, "monto_no_parseable", "El monto informado no pudo convertirse a número.", "media")
                if product == "Sin dato":
                    add_alert(alerts, record, "producto_no_inferido", "No se identificó producto o insumo en columna, hoja o archivo.", "media")

            processed_sheets.append({
                "archivo": filename, "hoja": sheetname, "filas_normalizadas": len(records) - before,
                "encabezado_fila": header_row,
            })
        workbook.close()

    fact = pd.DataFrame(records, columns=FACT_COLUMNS)
    if not fact.empty:
        identity = fact["cuit_cuil_norm"].where(fact["cuit_cuil_norm"].ne(""), fact["documento_norm"])
        identity = identity.where(identity.ne(""), fact["productor_nombre"].map(normalize_text))
        duplicate_columns = ["anio", "proveedor", "insumo_producto", "cantidad", "monto_estimado"]
        duplicate_frame = pd.concat([identity.rename("identidad"), fact[duplicate_columns]], axis=1)
        duplicated = identity.ne("") & duplicate_frame.duplicated(keep=False)
        for index in fact.index[duplicated]:
            add_alert(
                alerts, fact.loc[index].to_dict(), "posible_duplicado",
                "Coincide con otra fila en identificación, año, proveedor, producto, cantidad y monto; no se eliminó.",
                "media",
            )
    quality = pd.DataFrame(alerts, columns=ALERT_COLUMNS)
    return fact, quality, processed_sheets


def write_summary(fact: pd.DataFrame, quality: pd.DataFrame, processed_sheets: list[dict[str, Any]]) -> None:
    years = sorted(str(int(value)) for value in fact["anio"].dropna().unique())
    providers = sorted(value for value in fact["proveedor"].dropna().unique() if value)
    assistance = fact["tipo_asistencia"].value_counts(dropna=False)
    identification = fact["calidad_identificacion"].value_counts().reindex(["alta", "media", "baja", "pendiente"], fill_value=0)
    alert_counts = quality["alerta_tipo"].value_counts() if not quality.empty else pd.Series(dtype=int)
    duplicate_count = int((quality["alerta_tipo"] == "posible_duplicado").sum()) if not quality.empty else 0
    files = sorted(fact["fuente_archivo"].unique())

    lines = [
        "# Resumen de transformación de entregas por emergencias agropecuarias", "",
        f"- **Fecha de ejecución:** {datetime.now().astimezone().isoformat(timespec='seconds')}",
        f"- **Archivos procesados:** {len(files)}", f"- **Hojas procesadas:** {len(processed_sheets)}",
        f"- **Filas normalizadas:** {len(fact)}", f"- **Años detectados:** {', '.join(years) or 'Ninguno'}",
        "- **Alcance:** transformación local; sin carga a TiDB ni modificación de dashboard o vistas SQL.", "",
        "## Archivos y hojas procesados", "",
    ]
    for item in processed_sheets:
        lines.append(f"- `{item['archivo']}` / **{item['hoja']}**: {item['filas_normalizadas']} filas; encabezado fuente en fila {item['encabezado_fila']}.")
    lines.extend(["", "## Cobertura e identificación", "",
        f"- CUIT/CUIL válido: **{int(fact['cuit_cuil_norm'].str.len().eq(11).sum())}** filas.",
        f"- Documento: **{int(fact['documento_norm'].ne('').sum())}** filas.",
        f"- RENSPA: **{int(fact['renspa'].ne('').sum())}** filas.",
        f"- ADREMA: **{int(fact['adrema'].ne('').sum())}** filas.", "",
        "| Calidad de identificación | Filas |", "|---|---:|",
    ])
    for label, count in identification.items():
        lines.append(f"| {label} | {int(count)} |")
    lines.extend(["", "## Proveedores detectados", "", ", ".join(providers) or "Sin datos", "", "## Tipos de asistencia", "", "| Tipo | Filas |", "|---|---:|"])
    for label, count in assistance.items():
        lines.append(f"| {label or 'Sin dato'} | {int(count)} |")
    lines.extend(["", "## Alertas de calidad", "", "| Alerta | Filas |", "|---|---:|"])
    if alert_counts.empty:
        lines.append("| Sin alertas | 0 |")
    else:
        for label, count in alert_counts.items():
            lines.append(f"| {label} | {int(count)} |")
    lines.extend(["", f"- **Filas marcadas como posibles duplicados:** {duplicate_count}. No fueron eliminadas.", "",
        "## Recomendación de integración a Ficha Productor", "",
        "1. Vincular primero por `cuit_cuil_norm` exacto de 11 dígitos.",
        "2. Usar `documento_norm` junto con `productor_nombre` para registros sin CUIT/CUIL válido.",
        "3. Usar RENSPA o ADREMA con nombre y ubicación como claves agropecuarias/prediales complementarias.",
        "4. Enviar registros de calidad baja o pendiente a revisión manual; no vincular solo por similitud nominal sin controles.",
        "5. Revisar los posibles duplicados entre hojas operativas y hojas de rendición antes de una carga futura.",
        "6. Conservar `fuente_archivo`, `fuente_hoja`, `source_row_number` y `entrega_id` para trazabilidad.", "",
        "## Limitaciones", "",
        "- Las inferencias de producto y tipo de asistencia se basan en columnas, nombres de hoja y archivo; no reemplazan validación administrativa.",
        "- Un monto ausente no se interpreta como cero.",
        "- No se inventaron CUIT/CUIL, documentos, RENSPA, ADREMA, fechas, cantidades ni montos.",
        "- Las filas duplicadas se conservan porque pueden representar rendición, copia operativa o eventos distintos.",
    ])
    (OUTPUT_DIR / "transform_entregas_resumen.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


def main() -> None:
    required = [
        AUDIT_DIR / "entregas_excel_inventory.csv", AUDIT_DIR / "entregas_excel_sheets.csv",
        AUDIT_DIR / "entregas_excel_columns.csv", AUDIT_DIR / "entregas_excel_key_fields_detected.csv",
        AUDIT_DIR / "entregas_excel_audit_resultados.md",
    ]
    missing = [str(path) for path in required if not path.exists()]
    if missing:
        raise FileNotFoundError("Faltan insumos de auditoría:\n- " + "\n- ".join(missing))
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    fact, quality, processed_sheets = transform()
    fact.to_csv(OUTPUT_DIR / "fact_entregas_emergencia.csv", index=False, encoding="utf-8-sig")
    quality.to_csv(OUTPUT_DIR / "fact_entregas_calidad_dato.csv", index=False, encoding="utf-8-sig")
    write_summary(fact, quality, processed_sheets)
    print(f"Transformación finalizada: {len(fact)} filas, {len(processed_sheets)} hojas, {len(quality)} alertas.")
    print(f"Salidas: {OUTPUT_DIR}")


if __name__ == "__main__":
    main()
