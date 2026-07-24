"""Auditoría reproducible del Excel de DDJJ/trámites agropecuarios.

El script es deliberadamente de solo lectura respecto de la fuente. Genera un
informe Markdown y CSV agregados en data_processed/ddjj_2023_excel/audit/.

Ejecución desde la raíz del repositorio:
    python scripts/13_audit_ddjj_2023_excel.py
"""

from __future__ import annotations

import argparse
import csv
import hashlib
import math
import re
import statistics
import unicodedata
from collections import Counter
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Iterable, Sequence

import openpyxl


PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_INPUT = (
    PROJECT_ROOT
    / "data_raw"
    / "ddjj_2023_excel"
    / "original"
    / "Informes Tramites Emerg. Agrop - Actualizado 21052026.xlsx"
)
DEFAULT_OUTPUT_DIR = (
    PROJECT_ROOT / "data_processed" / "ddjj_2023_excel" / "audit"
)

EXPECTED_LIVESTOCK_SHEETS = (
    "GANADERIA - Bovinos_Buf",
    "GANADERIA - Ovina",
    "GANADERIA - Porcinos",
    "GANADERIA - Equinos",
    "GANADERIA - Bovinos Datos Adic",
    "GANADERIA - Bovinos Nota Fucosa",
    "GANADERIA - Ovina Datos A",
    "GANADERIA - Porcinos Datos Adic",
    "GANADERIA - Equinos Datos Adic.",
)


@dataclass
class SheetData:
    name: str
    header_row: int | None
    headers: list[str | None]
    rows: list[list[Any]]
    max_row: int
    max_column: int


def is_blank(value: Any) -> bool:
    return value is None or (isinstance(value, str) and value.strip() == "")


def text_value(value: Any) -> str | None:
    if is_blank(value):
        return None
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).strip()


def canonical_name(value: Any) -> str:
    """Normaliza nombres solo para comparar; nunca reemplaza el nombre original."""
    text = unicodedata.normalize("NFKD", str(value or ""))
    text = "".join(char for char in text if not unicodedata.combining(char))
    return re.sub(r"[^a-z0-9]", "", text.lower())


def normalize_key(value: Any) -> str | None:
    text = text_value(value)
    if text is None:
        return None
    return re.sub(r"\.0$", "", text).strip()


def find_column(headers: Sequence[str | None], candidates: Iterable[str]) -> int | None:
    expected = {canonical_name(candidate) for candidate in candidates}
    for index, header in enumerate(headers):
        if canonical_name(header) in expected:
            return index
    return None


def find_key_column(headers: Sequence[str | None]) -> int | None:
    """Reconoce Tramite Id, tramiteId, TRAMITE_ID y variantes equivalentes."""
    exact = find_column(headers, ("Tramite Id", "tramiteId", "TRAMITE_ID"))
    if exact is not None:
        return exact
    for index, header in enumerate(headers):
        normalized = canonical_name(header)
        if "tramite" in normalized and normalized.endswith("id"):
            return index
    return None


def get_cell(row: Sequence[Any], index: int | None) -> Any:
    if index is None or index >= len(row):
        return None
    return row[index]


def parse_number(value: Any) -> float | None:
    if is_blank(value):
        return None
    if isinstance(value, bool):
        return None
    if isinstance(value, (int, float)):
        number = float(value)
        return number if math.isfinite(number) else None

    text = str(value).strip().replace(" ", "")
    if not text:
        return None
    if "," in text and "." in text:
        if text.rfind(",") > text.rfind("."):
            text = text.replace(".", "").replace(",", ".")
        else:
            text = text.replace(",", "")
    elif "," in text:
        text = text.replace(",", ".")
    try:
        number = float(text)
    except ValueError:
        return None
    return number if math.isfinite(number) else None


def numeric_quality(rows: Sequence[Sequence[Any]], column_index: int | None) -> dict[str, int]:
    if column_index is None:
        return {"present": 0, "null": 0, "zero": 0, "negative": 0, "invalid": 0}
    null = zero = negative = invalid = present = 0
    for row in rows:
        raw = get_cell(row, column_index)
        if is_blank(raw):
            null += 1
            continue
        present += 1
        number = parse_number(raw)
        if number is None:
            invalid += 1
        elif number == 0:
            zero += 1
        elif number < 0:
            negative += 1
    return {
        "present": present,
        "null": null,
        "zero": zero,
        "negative": negative,
        "invalid": invalid,
    }


def parse_year(value: Any) -> int | None:
    if isinstance(value, (datetime, date)):
        return value.year
    if is_blank(value):
        return None
    text = str(value).strip()
    match = re.search(r"(?:19|20)\d{2}", text)
    if match:
        return int(match.group(0))
    for fmt in ("%d/%m/%Y", "%Y-%m-%d", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).year
        except ValueError:
            continue
    return None


def is_valid_cuit(value: Any) -> bool:
    digits = re.sub(r"\D", "", text_value(value) or "")
    if len(digits) != 11:
        return False
    multipliers = (5, 4, 3, 2, 7, 6, 5, 4, 3, 2)
    check = 11 - sum(int(digits[i]) * multipliers[i] for i in range(10)) % 11
    if check == 11:
        check = 0
    elif check == 10:
        check = 9
    return check == int(digits[-1])


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def read_workbook(path: Path) -> list[SheetData]:
    workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
    result: list[SheetData] = []
    try:
        for worksheet in workbook.worksheets:
            raw_rows = list(worksheet.iter_rows(values_only=True))
            header_position: int | None = None
            headers: list[str | None] = []
            for position, row in enumerate(raw_rows[:20], start=1):
                if any(not is_blank(value) for value in row):
                    header_position = position
                    headers = [text_value(value) for value in row]
                    break

            data_rows: list[list[Any]] = []
            if header_position is not None:
                for row in raw_rows[header_position:]:
                    if any(not is_blank(value) for value in row):
                        padded = list(row) + [None] * max(0, len(headers) - len(row))
                        data_rows.append(padded[: len(headers)])

            result.append(
                SheetData(
                    name=worksheet.title,
                    header_row=header_position,
                    headers=headers,
                    rows=data_rows,
                    max_row=worksheet.max_row,
                    max_column=worksheet.max_column,
                )
            )
    finally:
        workbook.close()
    return result


def sheet_by_name(sheets: Sequence[SheetData], expected: str) -> SheetData | None:
    target = canonical_name(expected)
    return next((sheet for sheet in sheets if canonical_name(sheet.name) == target), None)


def duplicate_headers(headers: Sequence[str | None]) -> list[str]:
    grouped: dict[str, list[str]] = {}
    for header in headers:
        if is_blank(header):
            continue
        grouped.setdefault(canonical_name(header), []).append(str(header))
    return [" / ".join(values) for values in grouped.values() if len(values) > 1]


def fully_blank_columns(sheet: SheetData) -> list[str]:
    blank_columns: list[str] = []
    for index, header in enumerate(sheet.headers):
        if all(is_blank(get_cell(row, index)) for row in sheet.rows):
            blank_columns.append(header or f"<sin nombre #{index + 1}>")
    return blank_columns


def suspicious_columns(sheet: SheetData) -> list[str]:
    issues: list[str] = []
    duplicates = duplicate_headers(sheet.headers)
    if duplicates:
        issues.append("duplicadas: " + "; ".join(duplicates))
    whitespace = [
        str(header)
        for header in sheet.headers
        if isinstance(header, str) and header != header.strip()
    ]
    if whitespace:
        issues.append("espacios laterales: " + "; ".join(whitespace))
    if "equinosdatosadic" in canonical_name(sheet.name):
        mislabeled = [
            str(header)
            for header in sheet.headers
            if canonical_name(header).startswith("porcinos")
        ]
        if mislabeled:
            issues.append("rótulos PORCINOS_* en hoja de Equinos: " + "; ".join(mislabeled))
    return issues


def distribution(rows: Sequence[Sequence[Any]], index: int | None) -> Counter[str]:
    if index is None:
        return Counter()
    return Counter(text_value(get_cell(row, index)) or "<nulo>" for row in rows)


def coalesced_distribution(
    rows: Sequence[Sequence[Any]], indices: Sequence[int | None]
) -> Counter[str]:
    """Usa la primera categoría no vacía por fila respetando el orden indicado."""
    result: Counter[str] = Counter()
    valid_indices = [index for index in indices if index is not None]
    for row in rows:
        category = next(
            (
                value
                for index in valid_indices
                if (value := text_value(get_cell(row, index))) is not None
            ),
            "<nulo>",
        )
        result[category] += 1
    return result


def format_distribution(values: Counter[str], limit: int | None = None) -> str:
    items = values.most_common(limit)
    return "; ".join(f"{label}={count}" for label, count in items) or "sin datos"


def linkage_quality(sheet: SheetData, main_keys: set[str]) -> dict[str, Any]:
    key_index = find_key_column(sheet.headers)
    if key_index is None:
        return {
            "key_column": None,
            "key_null": len(sheet.rows),
            "key_unique": 0,
            "linked_unique": 0,
            "orphan_rows": 0,
            "orphan_unique": 0,
        }
    keys = [normalize_key(get_cell(row, key_index)) for row in sheet.rows]
    valid = {key for key in keys if key}
    orphan = {key for key in valid if key not in main_keys}
    return {
        "key_column": sheet.headers[key_index],
        "key_null": sum(key is None for key in keys),
        "key_unique": len(valid),
        "linked_unique": len(valid & main_keys),
        "orphan_rows": sum(key is not None and key not in main_keys for key in keys),
        "orphan_unique": len(orphan),
    }


def audit_inventory(sheets: Sequence[SheetData], main_keys: set[str]) -> list[dict[str, Any]]:
    rows: list[dict[str, Any]] = []
    for sheet in sheets:
        unnamed = [
            f"columna {index + 1}"
            for index, header in enumerate(sheet.headers)
            if is_blank(header)
        ]
        blank_columns = fully_blank_columns(sheet)
        duplicates = duplicate_headers(sheet.headers)
        suspicious = suspicious_columns(sheet)
        linkage = linkage_quality(sheet, main_keys)
        rows.append(
            {
                "sheet_name": sheet.name,
                "header_row": sheet.header_row or "",
                "excel_max_row": sheet.max_row,
                "excel_max_column": sheet.max_column,
                "data_rows": len(sheet.rows),
                "columns": len(sheet.headers),
                "key_column": linkage["key_column"] or "",
                "key_null": linkage["key_null"],
                "key_unique": linkage["key_unique"],
                "linked_unique": linkage["linked_unique"],
                "orphan_rows": linkage["orphan_rows"],
                "orphan_unique": linkage["orphan_unique"],
                "unnamed_columns": "; ".join(unnamed),
                "fully_blank_columns": "; ".join(blank_columns),
                "duplicate_columns": "; ".join(duplicates),
                "suspicious_columns": "; ".join(suspicious),
                "useful_status": "sin registros útiles" if not sheet.rows else "con datos",
            }
        )
    return rows


def metric(section: str, name: str, value: Any, details: str = "") -> dict[str, Any]:
    return {"section": section, "metric": name, "value": value, "details": details}


def audit_tramites(sheet: SheetData | None) -> tuple[list[dict[str, Any]], set[str], dict[str, Any]]:
    if sheet is None:
        missing = [metric("estructura", "hoja_Tramites_presente", 0, "No se encontró la hoja principal")]
        return missing, set(), {"missing": True}

    key_index = find_key_column(sheet.headers)
    cuit_index = find_column(sheet.headers, ("CUIT", "CUIT/CUIL", "cuit_cuil"))
    name_index = find_column(sheet.headers, ("Razón Social", "Razon Social", "productor_nombre"))
    date_index = find_column(sheet.headers, ("Fecha de Creación", "Fecha Creacion"))
    state_index = find_column(sheet.headers, ("Estado Actual", "estado"))
    expired_index = find_column(sheet.headers, ("Caduco", "caducidad"))
    certificate_type_index = find_column(sheet.headers, ("Tipo Certificado",))
    certificate_number_index = find_column(sheet.headers, ("Numero Certificado", "Número Certificado"))

    keys = [normalize_key(get_cell(row, key_index)) for row in sheet.rows]
    valid_keys = [key for key in keys if key]
    key_counts = Counter(valid_keys)
    main_keys = set(valid_keys)

    cuits = [text_value(get_cell(row, cuit_index)) for row in sheet.rows]
    valid_cuit_values = [value for value in cuits if value]
    years = Counter(parse_year(get_cell(row, date_index)) for row in sheet.rows)
    years.pop(None, None)
    states = distribution(sheet.rows, state_index)
    expired = distribution(sheet.rows, expired_index)
    certificate_types = distribution(sheet.rows, certificate_type_index)

    metrics = [
        metric("tramites", "filas_totales", len(sheet.rows)),
        metric("tramites", "tramite_id_columna", sheet.headers[key_index] if key_index is not None else "faltante"),
        metric("tramites", "tramite_id_nulos", sum(key is None for key in keys)),
        metric("tramites", "tramites_unicos", len(main_keys)),
        metric("tramites", "tramites_duplicados_filas_excedentes", sum(count - 1 for count in key_counts.values() if count > 1)),
        metric("tramites", "tramite_id_duplicados_distintos", sum(count > 1 for count in key_counts.values())),
        metric("productores", "cuit_unicos", len(set(valid_cuit_values))),
        metric("productores", "cuit_nulos", sum(value is None for value in cuits)),
        metric("productores", "cuit_invalidos", sum(value is not None and not is_valid_cuit(value) for value in cuits)),
        metric("productores", "razon_social_nula", sum(is_blank(get_cell(row, name_index)) for row in sheet.rows)),
        metric("fechas", "fecha_creacion_nula", sum(is_blank(get_cell(row, date_index)) for row in sheet.rows)),
        metric("fechas", "tramites_fuera_2023", sum((year := parse_year(get_cell(row, date_index))) is not None and year != 2023 for row in sheet.rows)),
        metric("fechas", "tramites_por_anio", len(sheet.rows), format_distribution(Counter({str(year): count for year, count in sorted(years.items())}))),
        metric("estado", "distribucion_estado_actual", len(sheet.rows), format_distribution(states)),
        metric("estado", "distribucion_caduco", len(sheet.rows), format_distribution(expired)),
        metric("certificado", "distribucion_tipo_certificado", len(sheet.rows), format_distribution(certificate_types)),
        metric("certificado", "numero_certificado_nulo", sum(is_blank(get_cell(row, certificate_number_index)) for row in sheet.rows)),
    ]

    summary = {
        "missing": False,
        "rows": len(sheet.rows),
        "unique": len(main_keys),
        "duplicate_rows": sum(count - 1 for count in key_counts.values() if count > 1),
        "cuit_unique": len(set(valid_cuit_values)),
        "cuit_null": sum(value is None for value in cuits),
        "cuit_invalid": sum(value is not None and not is_valid_cuit(value) for value in cuits),
        "name_null": sum(is_blank(get_cell(row, name_index)) for row in sheet.rows),
        "date_null": sum(is_blank(get_cell(row, date_index)) for row in sheet.rows),
        "outside_2023": sum((year := parse_year(get_cell(row, date_index))) is not None and year != 2023 for row in sheet.rows),
        "years": years,
        "states": states,
        "expired": expired,
        "certificate_types": certificate_types,
        "certificate_null": sum(is_blank(get_cell(row, certificate_number_index)) for row in sheet.rows),
    }
    return metrics, main_keys, summary


def audit_adremas(sheet: SheetData | None, main_keys: set[str]) -> tuple[list[dict[str, Any]], dict[str, Any]]:
    if sheet is None:
        rows = [metric("estructura", "hoja_ADREMAS_presente", 0, "No se encontró la hoja")]
        return rows, {"missing": True}

    linkage = linkage_quality(sheet, main_keys)
    key_index = find_key_column(sheet.headers)
    adrema_index = find_column(sheet.headers, ("adrema",))
    surface_index = find_column(sheet.headers, ("superficie",))
    department_index = find_column(sheet.headers, ("departamentoDescr", "departamento"))
    municipality_index = find_column(sheet.headers, ("municipioDesc", "municipio"))
    location_index = find_column(sheet.headers, ("parajeDesc", "paraje"))
    surface = numeric_quality(sheet.rows, surface_index)

    counts = Counter(
        normalize_key(get_cell(row, key_index))
        for row in sheet.rows
        if normalize_key(get_cell(row, key_index)) in main_keys
    )
    count_values = list(counts.values())
    distribution_counts = Counter(count_values)
    linked_keys = set(counts)

    adrema_pairs = Counter(
        (normalize_key(get_cell(row, key_index)), text_value(get_cell(row, adrema_index)))
        for row in sheet.rows
        if normalize_key(get_cell(row, key_index)) and text_value(get_cell(row, adrema_index))
    )
    duplicated_header_list = duplicate_headers(sheet.headers)

    metrics = [
        metric("adremas", "filas_totales", len(sheet.rows)),
        metric("adremas", "clave_tramite", linkage["key_column"] or "faltante"),
        metric("adremas", "tramites_vinculados", linkage["linked_unique"]),
        metric("adremas", "filas_huerfanas", linkage["orphan_rows"]),
        metric("adremas", "tramites_huerfanos", linkage["orphan_unique"]),
        metric("adremas", "tramites_sin_adrema", len(main_keys - linked_keys)),
        metric("adremas", "adrema_nula", sum(is_blank(get_cell(row, adrema_index)) for row in sheet.rows)),
        metric("adremas", "pares_tramite_adrema_duplicados", sum(count > 1 for count in adrema_pairs.values())),
        metric("adremas", "filas_excedentes_tramite_adrema", sum(count - 1 for count in adrema_pairs.values() if count > 1)),
        metric("adremas_por_tramite", "minimo", min(count_values) if count_values else 0),
        metric("adremas_por_tramite", "promedio", round(statistics.mean(count_values), 3) if count_values else 0),
        metric("adremas_por_tramite", "mediana", statistics.median(count_values) if count_values else 0),
        metric("adremas_por_tramite", "maximo", max(count_values) if count_values else 0),
        metric("adremas_por_tramite", "distribucion", len(count_values), format_distribution(Counter({str(k): v for k, v in sorted(distribution_counts.items())}))),
        metric("superficie", "superficie_nula", surface["null"]),
        metric("superficie", "superficie_cero", surface["zero"]),
        metric("superficie", "superficie_negativa", surface["negative"]),
        metric("superficie", "superficie_no_numerica", surface["invalid"]),
        metric("ubicacion", "departamento_nulo", sum(is_blank(get_cell(row, department_index)) for row in sheet.rows)),
        metric("ubicacion", "municipio_nulo", sum(is_blank(get_cell(row, municipality_index)) for row in sheet.rows)),
        metric("ubicacion", "paraje_nulo", sum(is_blank(get_cell(row, location_index)) for row in sheet.rows)),
        metric("estructura", "columnas_duplicadas", len(duplicated_header_list), "; ".join(duplicated_header_list)),
    ]
    summary = {
        "missing": False,
        "rows": len(sheet.rows),
        "linked": linkage["linked_unique"],
        "orphan_rows": linkage["orphan_rows"],
        "without_adrema": len(main_keys - linked_keys),
        "surface": surface,
        "department_null": sum(is_blank(get_cell(row, department_index)) for row in sheet.rows),
        "municipality_null": sum(is_blank(get_cell(row, municipality_index)) for row in sheet.rows),
        "location_null": sum(is_blank(get_cell(row, location_index)) for row in sheet.rows),
        "duplicates": duplicated_header_list,
    }
    return metrics, summary


def audit_production(sheet: SheetData | None, main_keys: set[str]) -> dict[str, Any]:
    if sheet is None:
        return {"missing": True, "metrics": [metric("estructura", "hoja_PERDIDAS_PRODUCCION_presente", 0)]}

    linkage = linkage_quality(sheet, main_keys)
    category_description_index = find_column(sheet.headers, ("especieDesc", "productoDesc", "rubroDesc", "cultivoDesc"))
    category_id_index = find_column(sheet.headers, ("especieId", "productoId", "rubroId", "cultivoId"))
    category_generic_index = find_column(sheet.headers, ("producto", "rubro", "cultivo", "especie"))
    category_indices = (category_description_index, category_generic_index, category_id_index)
    category_columns = [sheet.headers[index] for index in category_indices if index is not None]
    planted_index = find_column(sheet.headers, ("superficieSembrada", "superficie_sembrada"))
    affected_index = find_column(sheet.headers, ("superficieAfectada", "superficie_afectada"))
    planted = numeric_quality(sheet.rows, planted_index)
    affected = numeric_quality(sheet.rows, affected_index)
    categories = coalesced_distribution(sheet.rows, category_indices)
    metrics = [
        metric("produccion", "filas_totales", len(sheet.rows)),
        metric("produccion", "clave_tramite", linkage["key_column"] or "faltante"),
        metric("produccion", "tramites_vinculados", linkage["linked_unique"]),
        metric("produccion", "filas_huerfanas", linkage["orphan_rows"]),
        metric("produccion", "tramites_huerfanos", linkage["orphan_unique"]),
        metric("produccion", "columnas_categoria_priorizadas", "; ".join(category_columns) if category_columns else "faltante"),
        metric("produccion", "registros_por_categoria", len(sheet.rows), format_distribution(categories)),
        metric("superficie_sembrada", "nula", planted["null"]),
        metric("superficie_sembrada", "cero", planted["zero"]),
        metric("superficie_sembrada", "negativa", planted["negative"]),
        metric("superficie_sembrada", "no_numerica", planted["invalid"]),
        metric("superficie_afectada", "nula", affected["null"]),
        metric("superficie_afectada", "cero", affected["zero"]),
        metric("superficie_afectada", "negativa", affected["negative"]),
        metric("superficie_afectada", "no_numerica", affected["invalid"]),
    ]
    return {
        "missing": False,
        "metrics": metrics,
        "rows": len(sheet.rows),
        "linked": linkage["linked_unique"],
        "orphan_rows": linkage["orphan_rows"],
        "category_column": "; ".join(category_columns) if category_columns else None,
        "categories": categories,
        "planted": planted,
        "affected": affected,
    }


def audit_livestock(sheets: Sequence[SheetData], main_keys: set[str]) -> list[dict[str, Any]]:
    output: list[dict[str, Any]] = []
    for expected in EXPECTED_LIVESTOCK_SHEETS:
        sheet = sheet_by_name(sheets, expected)
        if sheet is None:
            output.append(
                {
                    "sheet_name": expected,
                    "sheet_present": 0,
                    "sheet_type": "adicional" if "datos" in canonical_name(expected) or "nota" in canonical_name(expected) else "principal",
                    "rows": 0,
                    "key_column": "",
                    "linked_unique": 0,
                    "orphan_rows": 0,
                    "orphan_unique": 0,
                    "category_column": "",
                    "categories": "",
                    "quantity_null": "",
                    "quantity_zero": "",
                    "quantity_negative": "",
                    "quantity_invalid": "",
                    "mortality_null": "",
                    "mortality_zero": "",
                    "mortality_negative": "",
                    "mortality_invalid": "",
                    "mortality_gt_quantity": "",
                    "surface_columns_quality": "",
                    "suspicious_columns": "hoja esperada faltante",
                }
            )
            continue

        normalized_sheet = canonical_name(sheet.name)
        sheet_type = "adicional" if "datos" in normalized_sheet or "nota" in normalized_sheet else "principal"
        linkage = linkage_quality(sheet, main_keys)
        category_description_index = find_column(sheet.headers, ("descripcionAnimalDesc",))
        category_code_index = find_column(sheet.headers, ("descripcionAnimal",))
        category_generic_index = find_column(sheet.headers, ("categoria", "especie"))
        category_indices = (category_description_index, category_generic_index, category_code_index)
        category_columns = [sheet.headers[index] for index in category_indices if index is not None]
        quantity_index = find_column(sheet.headers, ("cantidad", "existencias"))
        mortality_index = find_column(sheet.headers, ("mortandad", "perdidas"))
        quantity = numeric_quality(sheet.rows, quantity_index)
        mortality = numeric_quality(sheet.rows, mortality_index)
        categories = coalesced_distribution(sheet.rows, category_indices)

        mortality_gt_quantity = 0
        if quantity_index is not None and mortality_index is not None:
            for row in sheet.rows:
                quantity_value = parse_number(get_cell(row, quantity_index))
                mortality_value = parse_number(get_cell(row, mortality_index))
                if quantity_value is not None and mortality_value is not None and mortality_value > quantity_value:
                    mortality_gt_quantity += 1

        surface_details: list[str] = []
        for index, header in enumerate(sheet.headers):
            normalized = canonical_name(header)
            if "superficie" in normalized or "supuso" in normalized or "supafectada" in normalized:
                quality = numeric_quality(sheet.rows, index)
                surface_details.append(
                    f"{header}: nulos={quality['null']}, cero={quality['zero']}, "
                    f"negativos={quality['negative']}, no_numéricos={quality['invalid']}"
                )

        output.append(
            {
                "sheet_name": sheet.name,
                "sheet_present": 1,
                "sheet_type": sheet_type,
                "rows": len(sheet.rows),
                "key_column": linkage["key_column"] or "",
                "linked_unique": linkage["linked_unique"],
                "orphan_rows": linkage["orphan_rows"],
                "orphan_unique": linkage["orphan_unique"],
                "category_column": "; ".join(category_columns),
                "categories": format_distribution(categories),
                "quantity_null": quantity["null"] if quantity_index is not None else "",
                "quantity_zero": quantity["zero"] if quantity_index is not None else "",
                "quantity_negative": quantity["negative"] if quantity_index is not None else "",
                "quantity_invalid": quantity["invalid"] if quantity_index is not None else "",
                "mortality_null": mortality["null"] if mortality_index is not None else "",
                "mortality_zero": mortality["zero"] if mortality_index is not None else "",
                "mortality_negative": mortality["negative"] if mortality_index is not None else "",
                "mortality_invalid": mortality["invalid"] if mortality_index is not None else "",
                "mortality_gt_quantity": mortality_gt_quantity if mortality_index is not None and quantity_index is not None else "",
                "surface_columns_quality": "; ".join(surface_details),
                "suspicious_columns": "; ".join(suspicious_columns(sheet)),
            }
        )
    return output


def write_csv(path: Path, rows: Sequence[dict[str, Any]]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not rows:
        path.write_text("", encoding="utf-8-sig")
        return
    with path.open("w", newline="", encoding="utf-8-sig") as handle:
        writer = csv.DictWriter(handle, fieldnames=list(rows[0].keys()), extrasaction="ignore")
        writer.writeheader()
        writer.writerows(rows)


def markdown_escape(value: Any) -> str:
    return str(value if value is not None else "").replace("|", "\\|").replace("\n", " ")


def markdown_table(headers: Sequence[str], rows: Sequence[Sequence[Any]]) -> str:
    lines = [
        "| " + " | ".join(markdown_escape(header) for header in headers) + " |",
        "| " + " | ".join("---" for _ in headers) + " |",
    ]
    lines.extend("| " + " | ".join(markdown_escape(value) for value in row) + " |" for row in rows)
    return "\n".join(lines)


def build_markdown(
    source: Path,
    audit_time: datetime,
    sha256: str,
    inventory: Sequence[dict[str, Any]],
    tramites_metrics: Sequence[dict[str, Any]],
    tramites_summary: dict[str, Any],
    adremas_metrics: Sequence[dict[str, Any]],
    adremas_summary: dict[str, Any],
    production: dict[str, Any],
    livestock: Sequence[dict[str, Any]],
) -> str:
    empty_sheets = [row["sheet_name"] for row in inventory if row["data_rows"] == 0]
    orphan_sheets = [
        f"{row['sheet_name']} ({row['orphan_rows']} filas)"
        for row in inventory
        if int(row["orphan_rows"] or 0) > 0
    ]
    livestock_main = [row for row in livestock if row["sheet_present"] == 1 and row["sheet_type"] == "principal"]
    mortality_alerts = sum(int(row["mortality_gt_quantity"] or 0) for row in livestock_main)
    equine_mislabeled = [
        row for row in livestock if "equinosdatosadic" in canonical_name(row["sheet_name"]) and "PORCINOS" in row["suspicious_columns"]
    ]

    alerts: list[str] = []
    if tramites_summary.get("missing"):
        alerts.append("No se encontró la hoja principal `Tramites`; las relaciones no pudieron validarse.")
    else:
        if tramites_summary["cuit_invalid"]:
            alerts.append(f"Se detectaron {tramites_summary['cuit_invalid']} CUIT inválidos.")
        if tramites_summary["outside_2023"]:
            alerts.append(f"Se detectaron {tramites_summary['outside_2023']} trámites con fecha de creación fuera de 2023.")
        if tramites_summary["duplicate_rows"]:
            alerts.append(f"Existen {tramites_summary['duplicate_rows']} filas excedentes por Tramite Id duplicado.")
    if orphan_sheets:
        alerts.append("Hay claves de trámite huérfanas en: " + ", ".join(orphan_sheets) + ".")
    if not adremas_summary.get("missing"):
        if adremas_summary["surface"]["negative"]:
            alerts.append(f"ADREMAS contiene {adremas_summary['surface']['negative']} superficies negativas.")
        if adremas_summary["without_adrema"]:
            alerts.append(f"Hay {adremas_summary['without_adrema']} trámites sin registros en ADREMAS.")
    if mortality_alerts:
        alerts.append(f"Hay {mortality_alerts} registros ganaderos con mortandad mayor que cantidad.")
    if equine_mislabeled:
        alerts.append("La hoja de datos adicionales de Equinos contiene columnas rotuladas como `PORCINOS_*`.")
    if empty_sheets:
        alerts.append("Hojas sin registros útiles: " + ", ".join(empty_sheets) + ".")
    if not alerts:
        alerts.append("No se detectaron alertas críticas con las reglas implementadas.")

    inventory_rows = [
        (
            row["sheet_name"], row["data_rows"], row["columns"], row["key_column"] or "—",
            row["orphan_rows"], row["unnamed_columns"] or "—", row["fully_blank_columns"] or "—",
            row["duplicate_columns"] or "—", row["suspicious_columns"] or "—",
        )
        for row in inventory
    ]
    livestock_rows = [
        (
            row["sheet_name"], "sí" if row["sheet_present"] else "no", row["sheet_type"], row["rows"],
            row["key_column"] or "—", row["linked_unique"], row["orphan_rows"], row["quantity_null"] if row["quantity_null"] != "" else "—",
            row["quantity_zero"] if row["quantity_zero"] != "" else "—", row["quantity_negative"] if row["quantity_negative"] != "" else "—",
            row["mortality_null"] if row["mortality_null"] != "" else "—", row["mortality_zero"] if row["mortality_zero"] != "" else "—",
            row["mortality_negative"] if row["mortality_negative"] != "" else "—", row["mortality_gt_quantity"] if row["mortality_gt_quantity"] != "" else "—",
        )
        for row in livestock
    ]

    lines = [
        "# Auditoría DDJJ/trámites agropecuarios",
        "",
        "## Identificación de la ejecución",
        "",
        f"- **Archivo:** `{source.name}`",
        f"- **Ruta relativa:** `{source.relative_to(PROJECT_ROOT) if source.is_relative_to(PROJECT_ROOT) else source}`",
        f"- **Tamaño:** {source.stat().st_size} bytes",
        f"- **SHA256:** `{sha256}`",
        f"- **Fecha/hora de auditoría:** {audit_time.isoformat(timespec='seconds')}",
        "- **Modo de lectura:** solo lectura (`openpyxl`, `read_only=True`, `data_only=True`).",
        "",
        "## Resumen ejecutivo",
        "",
        f"- Hojas encontradas: **{len(inventory)}**.",
        f"- Hojas sin registros útiles: **{len(empty_sheets)}**.",
        f"- Trámites totales: **{tramites_summary.get('rows', 0)}**.",
        f"- Trámites únicos: **{tramites_summary.get('unique', 0)}**.",
        f"- CUIT únicos: **{tramites_summary.get('cuit_unique', 0)}**.",
        f"- Trámites vinculados a ADREMAS: **{adremas_summary.get('linked', 0)}**.",
        f"- Registros de pérdidas de producción: **{production.get('rows', 0)}**.",
        f"- Casos ganaderos con mortandad mayor que cantidad: **{mortality_alerts}**.",
        "",
        "## Principales alertas",
        "",
        *[f"- {alert}" for alert in alerts],
        "",
        "## Inventario de hojas",
        "",
        markdown_table(
            ("Hoja", "Filas", "Columnas", "Clave", "Huérfanas", "Sin nombre", "Completamente vacías", "Duplicadas", "Sospechosas"),
            inventory_rows,
        ),
        "",
        "## Calidad de la hoja Tramites",
        "",
        markdown_table(
            ("Sección", "Métrica", "Valor", "Detalle"),
            [(row["section"], row["metric"], row["value"], row["details"] or "—") for row in tramites_metrics],
        ),
        "",
        "## Calidad de ADREMAS",
        "",
        markdown_table(
            ("Sección", "Métrica", "Valor", "Detalle"),
            [(row["section"], row["metric"], row["value"], row["details"] or "—") for row in adremas_metrics],
        ),
        "",
        "## Pérdidas de producción",
        "",
        markdown_table(
            ("Sección", "Métrica", "Valor", "Detalle"),
            [(row["section"], row["metric"], row["value"], row["details"] or "—") for row in production["metrics"]],
        ),
        "",
        "## Ganadería",
        "",
        markdown_table(
            ("Hoja", "Presente", "Tipo", "Filas", "Clave", "Trámites", "Huérfanas", "Cant. nula", "Cant. cero", "Cant. negativa", "Mort. nula", "Mort. cero", "Mort. negativa", "Mort. > cant."),
            livestock_rows,
        ),
        "",
        "### Categorías y controles adicionales",
        "",
    ]
    for row in livestock:
        lines.extend(
            [
                f"- **{row['sheet_name']}**: categorías: {row['categories'] or 'no aplicable'}; "
                f"superficies: {row['surface_columns_quality'] or 'no aplicable'}; "
                f"columnas sospechosas: {row['suspicious_columns'] or 'ninguna'}.",
            ]
        )

    lines.extend(
        [
            "",
            "## Notas metodológicas",
            "",
            "- `Tramite Id`, `tramiteId`, `TRAMITE_ID` y variantes equivalentes se comparan mediante una normalización interna; los nombres originales se conservan en los reportes.",
            "- Una fila huérfana es aquella cuya clave no aparece en la hoja principal `Tramites`.",
            "- Los números almacenados como texto se interpretan aceptando coma o punto decimal; los valores originales no se modifican.",
            "- `mortandad > cantidad` es una alerta de consistencia, no una regla automática de exclusión. Debe validarse el significado institucional de `cantidad`.",
            "- Los CSV contienen métricas agregadas y no reproducen CUIT ni razones sociales individuales.",
            "",
        ]
    )
    return "\n".join(lines)


def parse_args() -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Auditar el Excel de DDJJ/trámites agropecuarios.")
    parser.add_argument("--input", type=Path, default=DEFAULT_INPUT, help="Ruta del Excel fuente.")
    parser.add_argument("--output-dir", type=Path, default=DEFAULT_OUTPUT_DIR, help="Directorio de reportes.")
    return parser.parse_args()


def main() -> int:
    args = parse_args()
    source = args.input if args.input.is_absolute() else PROJECT_ROOT / args.input
    output_dir = args.output_dir if args.output_dir.is_absolute() else PROJECT_ROOT / args.output_dir

    if not source.exists():
        raise FileNotFoundError(f"No existe el archivo fuente: {source}")
    if not source.is_file():
        raise ValueError(f"La ruta fuente no es un archivo: {source}")

    audit_time = datetime.now().astimezone()
    source_hash = sha256_file(source)
    sheets = read_workbook(source)
    tramites_sheet = sheet_by_name(sheets, "Tramites")
    tramites_metrics, main_keys, tramites_summary = audit_tramites(tramites_sheet)

    inventory = audit_inventory(sheets, main_keys)
    adremas_metrics, adremas_summary = audit_adremas(sheet_by_name(sheets, "ADREMAS"), main_keys)
    production = audit_production(sheet_by_name(sheets, "PERDIDAS_PRODUCCION"), main_keys)
    livestock = audit_livestock(sheets, main_keys)

    output_dir.mkdir(parents=True, exist_ok=True)
    write_csv(output_dir / "sheets_inventory.csv", inventory)
    write_csv(output_dir / "tramites_quality.csv", tramites_metrics)
    write_csv(output_dir / "adremas_quality.csv", adremas_metrics)
    write_csv(output_dir / "ganaderia_quality.csv", livestock)

    report = build_markdown(
        source=source,
        audit_time=audit_time,
        sha256=source_hash,
        inventory=inventory,
        tramites_metrics=tramites_metrics,
        tramites_summary=tramites_summary,
        adremas_metrics=adremas_metrics,
        adremas_summary=adremas_summary,
        production=production,
        livestock=livestock,
    )
    report_path = output_dir / "audit_ddjj_2023_resultados.md"
    report_path.write_text(report, encoding="utf-8")

    print(f"Fuente leída correctamente: {source}")
    print(f"SHA256: {source_hash}")
    print(f"Hojas auditadas: {len(sheets)}")
    print(f"Trámites: {tramites_summary.get('rows', 0)}")
    print(f"Reporte: {report_path}")
    print("CSV generados: sheets_inventory.csv, tramites_quality.csv, adremas_quality.csv, ganaderia_quality.csv")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
