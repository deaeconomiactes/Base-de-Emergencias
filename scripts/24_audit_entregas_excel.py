"""Auditoría no destructiva de planillas de entregas por emergencias agropecuarias.

El script inventaría archivos y hojas, infiere encabezados y tipos, detecta campos
candidatos para Ficha Productor y genera únicamente salidas de auditoría.
No modifica los Excel fuente ni produce una base normalizada.
"""

from __future__ import annotations

import hashlib
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

import pandas as pd
from openpyxl import load_workbook


PROJECT_DIR = Path(__file__).resolve().parents[1]
SOURCE_DIR = Path(r"C:\Users\Usuario\Documents\BAUTI\Emergencias\Entregas")
OUTPUT_DIR = PROJECT_DIR / "data_processed" / "entregas" / "audit"
VALID_EXTENSIONS = {".xlsx", ".xls", ".xlsm"}
PREVIEW_LIMIT = 20
SCAN_ROW_LIMIT = 10_000
HEADER_SCAN_LIMIT = 30
TYPE_SAMPLE_LIMIT = 500

FIELD_PATTERNS: dict[str, tuple[str, ...]] = {
    "cuit": (r"\bcuit\b",),
    "cuil": (r"\bcuil\b",),
    "cuit_cuil": (r"cuit.*cuil", r"cuil.*cuit", r"cuitcuil", r"cuit_cuil"),
    "dni": (r"\bdni\b",),
    "documento": (r"document", r"nro.*doc", r"n.*doc"),
    "productor": (r"productor", r"titular"),
    "beneficiario": (r"beneficiari",),
    "razon_social": (r"razon.*social",),
    "nombre_apellido": (r"nombre.*apellido", r"apellido.*nombre", r"nombre.*completo"),
    "renspa": (r"renspa",),
    "adrema": (r"adrema",),
    "establecimiento": (r"establecimiento",),
    "campo": (r"\bcampo\b", r"predio"),
    "departamento": (r"departamento", r"\bdpto\b",),
    "localidad": (r"localidad",),
    "municipio": (r"municipio",),
    "paraje": (r"paraje",),
    "fecha_entrega": (r"fecha.*entrega", r"entrega.*fecha"),
    "fecha": (r"\bfecha\b",),
    "anio": (r"\banio\b", r"\bano\b"),
    "tipo_asistencia": (r"tipo.*asistencia",),
    "tipo_entrega": (r"tipo.*entrega",),
    "insumo": (r"insumo",),
    "producto": (r"producto", r"material"),
    "proveedor": (r"proveedor",),
    "cantidad": (r"cantidad", r"\bcant\b", r"unidades"),
    "unidad": (r"unidad", r"medida"),
    "monto": (r"monto",),
    "valor": (r"valor", r"precio"),
    "importe": (r"importe",),
    "subsidio": (r"subsidio",),
    "saldo": (r"saldo",),
    "expediente": (r"expediente", r"\bexpte\b", r"\bexp\b"),
    "resolucion": (r"resolucion", r"\bresol\b"),
    "decreto": (r"decreto",),
    "norma": (r"\bnorma\b", r"instrumento.*legal"),
    "observaciones": (r"observ", r"comentario",),
}

IDENTIFICATION_FIELDS = {
    "cuit", "cuil", "cuit_cuil", "dni", "documento", "productor",
    "beneficiario", "razon_social", "nombre_apellido",
}


def normalize_text(value: Any) -> str:
    if value is None:
        return ""
    text = unicodedata.normalize("NFKD", str(value))
    text = "".join(char for char in text if not unicodedata.combining(char))
    text = text.lower().strip()
    text = re.sub(r"[^a-z0-9]+", "_", text).strip("_")
    return text


def display_value(value: Any, max_length: int = 120) -> str:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        text = value.isoformat()
    else:
        text = str(value).strip()
    return text[:max_length]


def sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        for chunk in iter(lambda: source.read(1024 * 1024), b""):
            digest.update(chunk)
    return digest.hexdigest()


def infer_year(path: Path) -> str:
    match = re.search(r"\b(2022|2024|2025)\b", path.name)
    if match:
        return match.group(1)
    for part in reversed(path.parts):
        if part in {"2022", "2024", "2025"}:
            return part
    return ""


def infer_provider(filename: str) -> str:
    normalized = normalize_text(filename)
    providers = {
        "prieto": "Prieto", "kofman": "Kofman", "koffman": "Kofman",
        "fitosan": "Fitosan", "borderes": "Borderes",
    }
    found = [label for token, label in providers.items() if token in normalized]
    return found[0] if found else "otro"


def infer_source_type(filename: str) -> str:
    normalized = normalize_text(filename)
    if "saldo" in normalized and "emergencia" in normalized:
        return "saldo_emergencia"
    if "listado" in normalized and ("asistid" in normalized or "productor" in normalized):
        return "listado_asistidos"
    if "subsid" in normalized:
        return "subsidio"
    if "especie" in normalized:
        return "asistencia_especie"
    if "entrega" in normalized or "insumo" in normalized:
        return "entrega_insumos"
    return "otro"


def get_excel_files() -> list[Path]:
    return sorted(
        path for path in SOURCE_DIR.rglob("*")
        if path.is_file()
        and path.suffix.lower() in VALID_EXTENSIONS
        and not path.name.startswith("~$")
    )


def cell_is_nonempty(value: Any) -> bool:
    return value is not None and str(value).strip() != ""


def row_values(ws: Any, row_number: int, max_column: int) -> list[Any]:
    return [ws.cell(row=row_number, column=column).value for column in range(1, max_column + 1)]


def header_score(values: list[Any]) -> float:
    nonempty = [value for value in values if cell_is_nonempty(value)]
    if not nonempty:
        return -1
    texts = [display_value(value) for value in nonempty]
    textual = sum(not looks_numeric(text) for text in texts)
    keywords = sum(bool(detect_fields(text)) for text in texts)
    unique_ratio = len(set(texts)) / len(texts)
    return len(nonempty) + textual * 0.7 + keywords * 3 + unique_ratio


def detect_header_row(ws: Any, max_column: int) -> int | None:
    best_row, best_score = None, -1.0
    for row_number in range(1, min(ws.max_row, HEADER_SCAN_LIMIT) + 1):
        values = row_values(ws, row_number, max_column)
        score = header_score(values)
        if score > best_score:
            best_row, best_score = row_number, score
    return best_row if best_score >= 2 else None


def detect_header_row_from_matrix(matrix: list[list[Any]]) -> int | None:
    best_row, best_score = None, -1.0
    for row_number, values in enumerate(matrix[:HEADER_SCAN_LIMIT], start=1):
        score = header_score(values)
        if score > best_score:
            best_row, best_score = row_number, score
    return best_row if best_score >= 2 else None


def make_headers(values: list[Any]) -> list[str]:
    used: Counter[str] = Counter()
    headers: list[str] = []
    for position, value in enumerate(values, start=1):
        base = display_value(value) or f"sin_nombre_{position}"
        used[base] += 1
        headers.append(base if used[base] == 1 else f"{base}_{used[base]}")
    return headers


def detect_fields(header: Any) -> list[str]:
    normalized = normalize_text(header)
    matches: list[str] = []
    for field, patterns in FIELD_PATTERNS.items():
        if any(re.search(pattern, normalized) for pattern in patterns):
            matches.append(field)
    # Los campos compuestos prevalecen sobre sus componentes genéricos.
    if "cuit_cuil" in matches:
        matches = [field for field in matches if field not in {"cuit", "cuil"}]
    if "fecha_entrega" in matches and "fecha" in matches:
        matches.remove("fecha")
    return matches


def looks_numeric(value: Any) -> bool:
    if isinstance(value, (int, float)) and not isinstance(value, bool):
        return True
    text = str(value).strip()
    if not text:
        return False
    cleaned = re.sub(r"[$%\s]", "", text)
    if re.fullmatch(r"[-+]?\d{1,3}(?:\.\d{3})*(?:,\d+)?", cleaned):
        return True
    if re.fullmatch(r"[-+]?\d+(?:[.,]\d+)?", cleaned):
        return True
    return False


def looks_date(value: Any) -> bool:
    if isinstance(value, (datetime, date)):
        return True
    text = str(value).strip()
    return bool(
        re.fullmatch(r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}", text)
        or re.fullmatch(r"\d{4}[/-]\d{1,2}[/-]\d{1,2}", text)
    )


def approximate_type(values: list[Any]) -> str:
    nonempty = [value for value in values if cell_is_nonempty(value)]
    if not nonempty:
        return "vacío"
    kinds: set[str] = set()
    for value in nonempty[:TYPE_SAMPLE_LIMIT]:
        if looks_date(value):
            kinds.add("fecha")
        elif looks_numeric(value):
            kinds.add("numero")
        else:
            kinds.add("texto")
    return next(iter(kinds)) if len(kinds) == 1 else "mixto"


def quality_for(fields: set[str]) -> str:
    strong = bool(fields & {"cuit", "cuil", "cuit_cuil"})
    medium_ids = bool(fields & {"dni", "documento", "renspa", "adrema"})
    has_name = bool(fields & {"productor", "beneficiario", "razon_social", "nombre_apellido"})
    if strong or (medium_ids and has_name):
        return "alta"
    if medium_ids or has_name:
        return "media"
    if fields:
        return "baja"
    return "pendiente"


def main() -> None:
    OUTPUT_DIR.mkdir(parents=True, exist_ok=True)
    files = get_excel_files()
    inventory: list[dict[str, Any]] = []
    sheets: list[dict[str, Any]] = []
    columns: list[dict[str, Any]] = []
    previews: list[dict[str, Any]] = []
    key_fields: list[dict[str, Any]] = []
    report_details: list[dict[str, Any]] = []

    for path in files:
        file_observations: list[str] = []
        try:
            workbook = load_workbook(path, read_only=True, data_only=True, keep_links=False)
        except Exception as error:
            inventory.append({
                "nombre_archivo": path.name, "ruta_archivo": str(path),
                "extensión": path.suffix.lower(), "tamaño_bytes": path.stat().st_size,
                "fecha_modificación": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
                "hash_sha256": sha256_file(path), "cantidad_hojas": 0,
                "año_inferido": infer_year(path), "proveedor_referencia_inferida": infer_provider(path.name),
                "posible_tipo_fuente": infer_source_type(path.name),
                "observaciones": f"No se pudo abrir: {type(error).__name__}: {error}",
            })
            continue

        inventory.append({
            "nombre_archivo": path.name, "ruta_archivo": str(path),
            "extensión": path.suffix.lower(), "tamaño_bytes": path.stat().st_size,
            "fecha_modificación": datetime.fromtimestamp(path.stat().st_mtime).isoformat(timespec="seconds"),
            "hash_sha256": sha256_file(path), "cantidad_hojas": len(workbook.sheetnames),
            "año_inferido": infer_year(path), "proveedor_referencia_inferida": infer_provider(path.name),
            "posible_tipo_fuente": infer_source_type(path.name),
            "observaciones": "; ".join(file_observations),
        })

        for ws in workbook.worksheets:
            max_row = min(ws.max_row or 0, SCAN_ROW_LIMIT)
            max_col = ws.max_column or 0
            matrix = [
                list(row) for row in ws.iter_rows(
                    min_row=1, max_row=max_row, max_col=max_col, values_only=True
                )
            ] if max_row and max_col else []
            header_row = detect_header_row_from_matrix(matrix) if matrix else None
            headers = make_headers(matrix[header_row - 1]) if header_row else []
            data_start = (header_row + 1) if header_row else 1
            column_values: dict[int, list[Any]] = defaultdict(list)
            nonempty_rows = 0
            nonempty_column_positions: set[int] = set()
            duplicate_signatures: Counter[tuple[str, ...]] = Counter()
            preview_count = 0

            for row_number in range(data_start, max_row + 1):
                values = matrix[row_number - 1]
                if not any(cell_is_nonempty(value) for value in values):
                    continue
                nonempty_rows += 1
                signature = tuple(display_value(value, 500) for value in values)
                duplicate_signatures[signature] += 1
                for position, value in enumerate(values, start=1):
                    if cell_is_nonempty(value):
                        nonempty_column_positions.add(position)
                    if len(column_values[position]) < TYPE_SAMPLE_LIMIT:
                        column_values[position].append(value)

                if preview_count < PREVIEW_LIMIT and headers:
                    record: dict[str, Any] = {
                        "nombre_archivo": path.name, "nombre_hoja": ws.title,
                        "source_row_number": row_number,
                    }
                    # Conserva las columnas originales con prefijo para evitar colisiones entre hojas.
                    for position, (header, value) in enumerate(zip(headers, values), start=1):
                        if cell_is_nonempty(value):
                            record[f"original__{normalize_text(header) or f'col_{position}'}"] = display_value(value, 500)
                        for field in detect_fields(header):
                            candidate_name = f"candidata__{field}"
                            if candidate_name not in record or not record[candidate_name]:
                                record[candidate_name] = display_value(value, 500)
                    previews.append(record)
                    preview_count += 1

            detected_fields: set[str] = set()
            unnamed_count = 0
            mixed_columns: list[str] = []
            text_dates: list[str] = []
            text_amounts: list[str] = []
            id_format_issues: list[str] = []
            for position, header in enumerate(headers, start=1):
                values = column_values[position]
                detected_type = approximate_type(values)
                fields = detect_fields(header)
                field_set = set(fields)
                detected_fields.update(fields)
                if header.startswith("sin_nombre_"):
                    unnamed_count += 1
                if detected_type == "mixto":
                    mixed_columns.append(header)
                if field_set & {"fecha", "fecha_entrega"} and detected_type in {"texto", "mixto"}:
                    text_dates.append(header)
                if field_set & {"monto", "valor", "importe", "subsidio", "saldo"} and detected_type in {"texto", "mixto"}:
                    text_amounts.append(header)
                if field_set & {"cuit", "cuil", "cuit_cuil", "dni", "documento"}:
                    texts = [display_value(value) for value in values if cell_is_nonempty(value)]
                    if any(re.search(r"[-.]|\.0$|^0", text) for text in texts):
                        id_format_issues.append(header)
                nonempty = sum(cell_is_nonempty(value) for value in values)
                denominator = max(nonempty_rows, 1)
                examples = [display_value(value) for value in values if cell_is_nonempty(value)][:5]
                columns.append({
                    "nombre_archivo": path.name, "nombre_hoja": ws.title,
                    "columna_original": header, "columna_normalizada": normalize_text(header),
                    "posición_columna": position, "tipo_detectado_aproximado": detected_type,
                    "porcentaje_nulos_aproximado": round((1 - nonempty / denominator) * 100, 2),
                    "ejemplos_valores": " | ".join(examples),
                    "campos_candidatos_detectados": " | ".join(fields),
                })

            likely_summary = nonempty_rows <= 5 or not headers or len(detected_fields) < 2
            sheet_observations: list[str] = []
            if header_row and header_row > 1:
                sheet_observations.append("encabezado no ubicado en primera fila; posible título o encabezado múltiple")
            if unnamed_count:
                sheet_observations.append(f"{unnamed_count} columnas sin nombre")
            if mixed_columns:
                sheet_observations.append(f"formatos mezclados en {len(mixed_columns)} columnas")
            duplicate_count = sum(count - 1 for count in duplicate_signatures.values() if count > 1)
            if duplicate_count:
                sheet_observations.append(f"{duplicate_count} posibles filas duplicadas exactas")
            if id_format_issues:
                sheet_observations.append("identificadores con guiones, puntos, ceros iniciales o .0")
            if text_dates:
                sheet_observations.append("fechas posiblemente almacenadas como texto o con formatos mezclados")
            if text_amounts:
                sheet_observations.append("montos posiblemente almacenados como texto o con formatos mezclados")
            if detected_fields & IDENTIFICATION_FIELDS == set() and detected_fields:
                sheet_observations.append("sin identificador nominal/documental detectado")
            elif not detected_fields & {"cuit", "cuil", "cuit_cuil", "dni", "documento", "renspa", "adrema"} and detected_fields & {"productor", "beneficiario", "razon_social", "nombre_apellido"}:
                sheet_observations.append("nombres de productores sin identificador estructurado")

            sheets.append({
                "nombre_archivo": path.name, "nombre_hoja": ws.title,
                "cantidad_filas": ws.max_row or 0, "cantidad_columnas": max_col,
                "filas_no_vacias_aprox": nonempty_rows,
                "columnas_no_vacias_aprox": len(nonempty_column_positions),
                "primera_fila_con_datos_probable": header_row or "",
                "parece_tabla_o_resumen": "resumen" if likely_summary else "tabla",
                "observaciones": "; ".join(sheet_observations),
            })

            has = lambda candidates: bool(detected_fields & set(candidates))
            key_row = {
                "nombre_archivo": path.name, "nombre_hoja": ws.title,
                "tiene_cuit": has(("cuit", "cuil", "cuit_cuil")),
                "tiene_dni": has(("dni", "documento")),
                "tiene_nombre_productor": has(("productor", "beneficiario", "razon_social", "nombre_apellido")),
                "tiene_renspa": has(("renspa",)), "tiene_adrema": has(("adrema",)),
                "tiene_departamento": has(("departamento",)), "tiene_localidad": has(("localidad",)),
                "tiene_fecha": has(("fecha", "fecha_entrega", "anio")),
                "tiene_producto_insumo": has(("producto", "insumo", "tipo_asistencia", "tipo_entrega")),
                "tiene_cantidad": has(("cantidad", "unidad")),
                "tiene_monto": has(("monto", "valor", "importe", "subsidio", "saldo")),
                "tiene_norma": has(("norma", "resolucion", "decreto", "expediente")),
                "calidad_probable_vinculacion": quality_for(detected_fields),
            }
            key_fields.append(key_row)
            report_details.append({
                "archivo": path.name, "hoja": ws.title, "header_row": header_row,
                "fields": sorted(detected_fields), "observaciones": sheet_observations,
                "filas": nonempty_rows, "columnas": len(nonempty_column_positions),
                "tipo": "resumen" if likely_summary else "tabla",
            })
        workbook.close()

    pd.DataFrame(inventory).to_csv(OUTPUT_DIR / "entregas_excel_inventory.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(sheets).to_csv(OUTPUT_DIR / "entregas_excel_sheets.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(columns).to_csv(OUTPUT_DIR / "entregas_excel_columns.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(previews).to_csv(OUTPUT_DIR / "entregas_excel_preview.csv", index=False, encoding="utf-8-sig")
    pd.DataFrame(key_fields).to_csv(OUTPUT_DIR / "entregas_excel_key_fields_detected.csv", index=False, encoding="utf-8-sig")
    write_report(inventory, sheets, key_fields, report_details)
    print(f"Auditoría finalizada: {len(inventory)} archivos, {len(sheets)} hojas.")
    print(f"Salidas: {OUTPUT_DIR}")


def write_report(
    inventory: list[dict[str, Any]],
    sheets: list[dict[str, Any]],
    key_fields: list[dict[str, Any]],
    details: list[dict[str, Any]],
) -> None:
    lines = [
        "# Auditoría de Excel de entregas y asistencia por emergencias agropecuarias",
        "",
        f"- **Fecha de ejecución:** {datetime.now().astimezone().isoformat(timespec='seconds')}",
        f"- **Carpeta auditada:** `{SOURCE_DIR}`",
        f"- **Archivos encontrados:** {len(inventory)}",
        f"- **Hojas auditadas:** {len(sheets)}",
        "- **Alcance:** inventario y diagnóstico estructural; sin integración, normalización ni modificación de fuentes.",
        "",
        "## Resumen por archivo",
        "",
        "| Archivo | Año | Proveedor/referencia | Tipo probable | Hojas |",
        "|---|---:|---|---|---:|",
    ]
    for row in inventory:
        lines.append(
            f"| {row['nombre_archivo']} | {row['año_inferido'] or 'No inferido'} | "
            f"{row['proveedor_referencia_inferida']} | {row['posible_tipo_fuente']} | {row['cantidad_hojas']} |"
        )

    detail_by_file: dict[str, list[dict[str, Any]]] = defaultdict(list)
    for detail in details:
        detail_by_file[detail["archivo"]].append(detail)
    lines.extend(["", "## Estructura detectada de cada base", ""])
    for filename, file_details in detail_by_file.items():
        lines.append(f"### {filename}")
        lines.append("")
        for detail in file_details:
            fields = ", ".join(f"`{field}`" for field in detail["fields"]) or "ninguna columna clave reconocida"
            observations = "; ".join(detail["observaciones"]) or "sin alertas heurísticas destacadas"
            lines.append(
                f"- **{detail['hoja']}**: {detail['tipo']}; encabezado probable en fila "
                f"{detail['header_row'] or 'no detectada'}; {detail['filas']} filas y "
                f"{detail['columnas']} columnas no vacías aproximadas. Campos: {fields}. "
                f"Observaciones: {observations}."
            )
        lines.append("")

    def any_true(field: str) -> bool:
        return any(bool(row[field]) for row in key_fields)

    lines.extend([
        "## Cobertura de columnas clave",
        "",
        f"- RENSPA: **{'sí' if any_true('tiene_renspa') else 'no'}**.",
        f"- CUIT/CUIL: **{'sí' if any_true('tiene_cuit') else 'no'}**.",
        f"- DNI/documento: **{'sí' if any_true('tiene_dni') else 'no'}**.",
        f"- ADREMA: **{'sí' if any_true('tiene_adrema') else 'no'}**.",
        f"- Montos: **{'sí' if any_true('tiene_monto') else 'no'}**; cantidades/unidades: **{'sí' if any_true('tiene_cantidad') else 'no'}**.",
        f"- Fechas/año: **{'sí' if any_true('tiene_fecha') else 'no'}**.",
        f"- Normas/expedientes: **{'sí' if any_true('tiene_norma') else 'no'}**.",
        "",
        "El detalle por archivo y hoja, incluida la calidad probable de vinculación, se encuentra en "
        "`entregas_excel_key_fields_detected.csv`.",
        "",
        "## Problemas detectados",
        "",
    ])
    issues = Counter(issue for detail in details for issue in detail["observaciones"])
    if issues:
        for issue, count in issues.most_common():
            lines.append(f"- {issue}: {count} hoja(s).")
    else:
        lines.append("- No se detectaron problemas mediante las reglas heurísticas aplicadas.")

    lines.extend([
        "",
        "Estas alertas son diagnósticas: requieren validación visual antes de transformar datos. "
        "La búsqueda incluyó encabezados múltiples/no iniciales, hojas resumen, columnas sin nombre, "
        "formatos mezclados, identificadores con puntuación/ceros/.0, fechas o montos como texto, "
        "duplicados exactos y nombres sin identificador.",
        "",
        "## Recomendación de clave de integración",
        "",
        "Orden sugerido de vinculación, conservando siempre la fuente y una clasificación de calidad:",
        "",
        "1. **CUIT/CUIL exacto**, después de una futura normalización controlada a 11 dígitos.",
        "2. **DNI/documento**, validado por longitud y acompañado por nombre o ubicación cuando sea posible.",
        "3. **RENSPA**, preservando sus componentes y verificando que identifique establecimiento/productor según la fuente.",
        "4. **ADREMA**, como clave predial complementaria, no necesariamente personal.",
        "5. **Nombre del productor**, únicamente como enlace probabilístico y nunca como clave única automática.",
        "6. **Combinación de claves** (nombre + documento/RENSPA/ADREMA + localidad/departamento) para resolver casos ambiguos.",
        "",
        "No se recomienda integrar registros solo por nombre sin revisión, debido a homónimos, variantes ortográficas "
        "y posibles diferencias entre titular, productor y beneficiario.",
        "",
        "## Modelo tentativo para integración en Ficha Productor",
        "",
        "Futura tabla normalizada propuesta: `fact_entregas_emergencia`.",
        "",
        "| Campo sugerido | Uso tentativo |",
        "|---|---|",
        "| entrega_id | Identificador técnico único de la entrega |",
        "| anio | Año de referencia de la asistencia |",
        "| fecha_entrega | Fecha efectiva o informada de entrega |",
        "| productor_nombre | Nombre o razón social según fuente |",
        "| cuit_cuil | Identificador tributario/laboral normalizado |",
        "| documento | DNI u otro documento normalizado |",
        "| renspa | Identificación agropecuaria |",
        "| adrema | Identificación predial provincial |",
        "| departamento | Departamento provincial |",
        "| localidad | Localidad o municipio informado |",
        "| tipo_asistencia | Categoría de asistencia |",
        "| proveedor | Proveedor o referencia de entrega |",
        "| insumo_producto | Bien, insumo o producto entregado |",
        "| cantidad | Magnitud entregada |",
        "| unidad | Unidad de medida |",
        "| monto_estimado | Valor monetario informado o estimado, claramente distinguido |",
        "| norma_evento | Resolución, decreto u otra norma/evento |",
        "| expediente | Expediente administrativo |",
        "| fuente_archivo | Archivo Excel de origen |",
        "| fuente_hoja | Hoja Excel de origen |",
        "| source_row_number | Número de fila original |",
        "| calidad_identificacion | Alta, media, baja o pendiente |",
        "| observaciones | Notas de fuente, calidad y transformación futura |",
        "",
        "### Limitaciones del modelo tentativo",
        "",
        "- La granularidad real (entrega, productor, expediente o lote de insumos) debe confirmarse por hoja.",
        "- Los montos deben distinguir valores observados de estimaciones futuras.",
        "- RENSPA y ADREMA pueden identificar unidades productivas o prediales, no necesariamente personas.",
        "- Esta auditoría no transforma ni valida definitivamente identificadores o importes.",
    ])
    (OUTPUT_DIR / "entregas_excel_audit_resultados.md").write_text("\n".join(lines) + "\n", encoding="utf-8")


if __name__ == "__main__":
    main()
